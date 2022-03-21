from fpdf import FPDF
from openpyxl import load_workbook

wb = load_workbook('./Эксперты Абилимпикс 2022.xlsx')
sheet = wb[wb.sheetnames[0]]

for i in range(2, sheet.max_row):
    last_name = sheet.cell(row=i, column=1).value.upper()
    first_name = sheet.cell(row=i, column=2).value.upper()
    middle_name = sheet.cell(row=i, column=3).value.upper()
    full_name = f'{last_name} {first_name} {middle_name}'
    category = sheet.cell(row=i, column=15).value.lower().strip()
    competition = sheet.cell(row=i, column=16).value.upper()


    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_font('FuturaBold', '',
                 r'C:\Users\admin\AppData\Local\Microsoft\Windows\Fonts\ofont.ru_Futura PT (1).ttf', uni=True)
    pdf.add_font('FuturaRegular', '',
                 r'C:\Users\admin\AppData\Local\Microsoft\Windows\Fonts\ofont.ru_Futura PT.ttf', uni=True)
    pdf.add_page()
    if category == 'эксперт':
        pdf.image('ДИПЛОМ_эксперта.png', 0, 0, 210)
    else:
        pdf.image('ДИПЛОМ_гл_эксперта.png', 0, 0, 210)
    pdf.set_font('FuturaBold', '', 19)
    pdf.cell(0, 155, ln=True)
    pdf.cell(0, 10, full_name, ln=True, align='C')

    pdf.set_font('FuturaRegular', '', 19)

    pdf.cell(0, 15, ln=True)

    pdf.cell(0, 8, 'КОМПЕТЕНЦИЯ:', ln=True, align='C')
    pdf.cell(0, 8, competition, ln=True, align='C')

    pdf.cell(0, 15, ln=True)

    pdf.set_font('FuturaRegular', '', 16)
    pdf.cell(0, 8, 'г. Якутск                                                     20 – 30 марта 2022 года', align='C')

    pdf.output(f'{category[:4].capitalize()} {competition[:10]} {full_name.title()}.pdf')
