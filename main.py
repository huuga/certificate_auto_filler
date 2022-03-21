from fpdf import FPDF
from openpyxl import load_workbook

wb = load_workbook('./Конкурсанты Абилимпикс 2022.xlsx')
sheet = wb[wb.sheetnames[0]]

for i in range(1, sheet.max_row):
    last_name = sheet.cell(row=i, column=1).value
    if last_name:
        last_name = last_name.upper()
    first_name = sheet.cell(row=i, column=2).value
    if first_name:
        first_name = first_name.upper()
    middle_name = sheet.cell(row=i, column=3).value
    if middle_name:
        middle_name = middle_name.upper()
    full_name = f'{last_name} {first_name} {middle_name}'
    institute = sheet.cell(row=i, column=24).value
    competition = sheet.cell(row=i, column=30).value
    if competition:
        competition = competition.upper()
    category = sheet.cell(row=i, column=31).value
    if category:
        category = category.lower()
    if category[0].lower() == 'ш':
        category += 'и'
    else:
        category += 'ы'

    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_font('FuturaBold', '',
                 r'C:\Users\admin\AppData\Local\Microsoft\Windows\Fonts\ofont.ru_Futura PT (1).ttf', uni=True)
    pdf.add_font('FuturaRegular', '',
                 r'C:\Users\admin\AppData\Local\Microsoft\Windows\Fonts\ofont.ru_Futura PT.ttf', uni=True)
    pdf.add_page()
    pdf.image('ДИПЛОМ_КОНКУРСАНТА.png', 0, 0, 210)

    pdf.set_font('FuturaBold', '', 19)
    pdf.cell(0, 150, ln=True)
    pdf.cell(0, 10, full_name, ln=True, align='C')

    pdf.set_font('FuturaRegular', '', 19)
    if institute:
        pdf.multi_cell(0, 6, institute, ln=True, align='C')
    else:
        pdf.cell(0, 10, ln=True)

    pdf.cell(0, 15, ln=True)

    pdf.cell(0, 8, competition, ln=True, align='C')
    pdf.cell(0, 8, f'({category})', ln=True, align='C')

    pdf.cell(0, 10, ln=True)

    pdf.set_font('FuturaRegular', '', 16)
    pdf.cell(0, 8, 'г. Якутск                                                     20 – 30 марта 2022 года', align='C')

    pdf.output(f'{competition[:10]} {full_name.title()}.pdf')
